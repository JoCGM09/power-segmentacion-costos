from fileinput import filename
import json
import requests

from openpyxl import load_workbook
from datetime import date

def select_instance ():
    print('Select Instance')
    import subprocess

    child = subprocess.Popen(['ibmcloud','resource','service-instances', '--output', 'json'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    retorno = child.stdout.read()
    retorno_json = json.loads(retorno)
    retorno_filtrado = []
    ind=0

    for x in retorno_json:
        if x['dashboard_url'] :
            power = "power-iaas" in x['dashboard_url']
            if power:
                retorno_filtrado.append(x)

    return retorno_filtrado

def exibe_menu (msg, menu_json,campo_exibir, campo_retorno, retorno_opcional01, retorno_opcional02):
    print('Exibe Menu')
    msg_retorno=[]
    menu = '0 - Sair'
    ind = 0
    sair = False
    for x in menu_json:
        ind = ind + 1        
        menu = menu + '\n' + str(ind) + ' - ' + x[campo_exibir] + ' - ' + x['region_id'] 

    print(menu)

    while sair==False:
        opcao = input('\n\n' + msg)
        if opcao.isdigit()==True:
            opcao_int =int(opcao)
            is_between = 0 <= opcao_int <= ind
            if is_between:
                sair=True
    
    if opcao_int==0:
        msg_retorno=[0]
    else:
        msg_retorno=[ 
                        opcao_int,
                        menu_json[opcao_int-1][campo_exibir],
                        menu_json[opcao_int-1][campo_retorno],
                        menu_json[opcao_int-1][retorno_opcional01],
                        menu_json[opcao_int-1][retorno_opcional02]
        ]

    return msg_retorno

def get_powervs(instancia_id, instancia_crn, filename, token,region_id):
    print('*************************************')
    print('Listando servidores virtuais...')

    url = 'https://' + region_id + '.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/'
    url=url + instancia_id + '/pvm-instances'
    print("**********")
    print("GET PVM")
    print(url)

    headers = {
        'Content-Type': 'application/json',
        'CRN':instancia_crn,
        'Authorization':'Bearer '+token 
    }

    r = requests.get(url, headers=headers)
    filename = 'output/' + filename + '-' + instancia_id + '.txt'
    
    with open(filename,'w') as outfile:
        outfile.write(str(r.text))

    return r.json()

def get_volumes(instancia_id, virtual_servers, instancia_crn, token, vsname,region_id):
    print('*************************************')
    print('Listando volumes...')

    url_origem = 'https://' + region_id+ '.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/' + instancia_id

    headers = {
        'Content-Type': 'application/json',
        'CRN':instancia_crn,
        'Authorization':'Bearer '+token 
    }

    for virtual_server in virtual_servers['pvmInstances']:
        print('Listando volumes do servidor virtual ' + virtual_server['serverName'])
        url = url_origem + '/pvm-instances/' + virtual_server['pvmInstanceID'] + '/volumes'
        filename = 'output/' + vsname + '-' + instancia_id + '-' + virtual_server['serverName'] + '-' + virtual_server['pvmInstanceID'] + '.txt'

        r = requests.get(url, headers=headers)
    
        with open(filename,'w') as outfile:
            outfile.write(str(r.text))
            outfile.close

    return True

def create_excel(instancia_name,instancia_id,dc_id):
    print('Gerando excel...')
    wb = load_workbook('output/Modelo.xlsx')
    ws = wb['Multiple LPAR Price Estimate']

    processor_types = {'capped':'C','shared':'S','dedicated':'D'}
    so_types = {'aix':'AIX','rhel':'BYO Linux'}

    file_source = 'output/' + instancia_name+'-'+ instancia_id+'.txt'
    with open(file_source, 'r') as openfile:
        obj = openfile.read()

    obj_json = json.loads(obj)

    linha=11

    for virtual_server in obj_json['pvmInstances']:
        linha=linha+1
        print(linha)
        celula_hostname='A'+str(linha)
        celula_lparqtde='B'+str(linha)
        celula_datacenter='C'+str(linha)
        celula_system='D'+str(linha)
        celula_processortype='E'+str(linha)
        celula_deseriredcores='F'+str(linha)
        celula_memory='G'+str(linha)
        celula_os='H'+str(linha)
        celula_tier1='I'+str(linha)
        celula_tier3='J'+str(linha)

        ws[celula_hostname]=virtual_server['serverName']
        ws[celula_lparqtde]=1
        ws[celula_datacenter]=dc_id
        ws[celula_system]=virtual_server['sysType']
        ws[celula_processortype]=processor_types[virtual_server['procType']]
        ws[celula_deseriredcores]=virtual_server['processors']
        ws[celula_memory]=virtual_server['memory']
        ws[celula_os]=so_types[virtual_server['osType']]
 
        volumes = get_volumes_size(instancia_name,instancia_id,virtual_server['serverName'],virtual_server['pvmInstanceID'])
       
        ws[celula_tier1]=volumes['tier1']
        ws[celula_tier3]=volumes['tier3']

    hoje=date.today()
    filename='output/Power Virtual Server Price Estimator V9k - ' + instancia_name+ ' - '+  str(hoje) + '.xlsx'
    wb.save(filename)

    print('Arquivo gerado: '+filename)
    return True

def get_volumes_size(instancia_nome, instancia_id, virtualserver_name, virtualserver_id):

    filename = 'output/'+instancia_nome+'-'+instancia_id+'-'+virtualserver_name+'-'+virtualserver_id+'.txt'
    volumes={
        'tier3':0,
        'tier1':0,
        'standard-legacy':0
    }
    with open(filename, 'r') as openfile:
        obj = openfile.read()

    obj_json = json.loads(obj)
   
    for volume in obj_json['volumes']:
        volumes[volume['diskType']] = volumes[volume['diskType']] + volume['size']
        
    return volumes

def getEvents(instancia_id, instancia_crn, token,region_id,instancia_name):
    print('getEvents')
   
    url = 'https://' + region_id +  '.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/'
    # url = 'https://eu-de.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/'
    url=url + instancia_id + '/events?time=1638327600'
    print("URL EVENTS:" + url)
    headers = {
        'Content-Type': 'application/json',
        'CRN':instancia_crn,
        'Authorization':'Bearer '+token 
    }

    r = requests.get(url, headers=headers)
    
    hoje=date.today()
    filename="output/events-" + instancia_name + "-" + str(hoje) + ".txt"

    with open(filename,'w') as outfile:
            outfile.write(str(r.text))
            outfile.close

    obj_json = json.loads(r.content)

    print("=========")

    print(obj_json)
  
    for evento in obj_json['events']:
        #url = 'https://' + region_id +  '.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/'
        url = 'https://eu-de.power-iaas.cloud.ibm.com/pcloud/v1/cloud-instances/'
        url=url + instancia_id + '/events/' + evento['eventID']
        r = requests.get(url, headers=headers)
        print(json.loads(r.content))

 
    return True

def verifyRegion(region_id):
    regions_replace= {
        "sao01":"sao",
        "eu-de-1":"eu-de",
        "eu-de-2":"eu-de",

    }
   
    need_replace = region_id in regions_replace

    if need_replace:
        region_retorno=regions_replace[region_id]
    else:
        region_retorno= region_id
    
    return region_retorno

def verifyDC(dc_id):
    dc_replace= {
        "eu-de-1":"fra04",
        "eu-de-2":"eu-de",
        "us-south":"DAL12",
        "us-east":"WDC04"
    }
   
    need_replace = dc_id in dc_replace

    if need_replace:
        dc_retorno=dc_replace[dc_id]
    else:
        dc_retorno= dc_id
    
    return dc_retorno